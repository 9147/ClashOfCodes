from django.core.management.base import BaseCommand
from MainApp.models import Team, Problem
from django.conf import settings
import openpyxl
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class Command(BaseCommand):
    help = 'Generate Excel file of teams that have made submissions along with submission details'

    def add_arguments(self, parser):
        # Add the base_url argument (this is the parameter that you will pass)
        parser.add_argument('--base_url', type=str, help='Base URL for media files')

    def handle(self, *args, **kwargs):
        # Get the base URL from the command arguments
        base_url = kwargs.get('base_url', "http://127.0.0.1:8000/")  # Use default if not provided

        # Get all teams that have made a submission with accepted status
        submitted_problems = Problem.objects.select_related('team').filter(status='Accepted')

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Teams with Submissions"

        # Add headers
        headers = [
            "Team ID", "Team Name", "Leader Name", "Leader Email", "Leader Contact",
            "Member 1", "Member 2", "Member 3", "City", "College", "State", "Country",
            "Submission Title", "Description", "Solution", "Domain", "Track", "Status",
            "Solution File (Link)","Github Link", "Payment Amount", "Payment UTR", "Payment Status", "Payment Date", "Payment Screenshot"
        ]
        sheet.append(headers)

        # Adjust column widths
        for i, header in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 20

        # Populate data
        for problem in submitted_problems:
            team = problem.team
            solution_url = f"{base_url}{settings.MEDIA_URL}{problem.solution_pdf}" if problem.solution_pdf else "No file uploaded"

            # Get payment details
            payment = getattr(team.leader, 'payment', None)  # Safely fetch payment related to the leader
            payment_amount = payment.amount if payment else "N/A"
            payment_utr = payment.utr_number if payment else "N/A"
            payment_status = payment.payment_status if payment else "N/A"
            payment_date = payment.posted_date.strftime('%Y-%m-%d %H:%M:%S') if payment else "N/A"
            payment_screenshot = f"{base_url}{settings.MEDIA_URL}{payment.payment_screenshot}" if payment and payment.payment_screenshot else "No screenshot uploaded"

            # Add the data
            row_data = [
                team.id, team.name, team.leader.first_name, team.leader.email, team.leader_contact,
                team.member1_name, team.member2_name, team.member3_name,
                team.city, team.college, team.state, team.country,
                problem.title, problem.description, problem.solution, problem.domain,
                problem.track, problem.status, solution_url, problem.github_link,
                payment_amount, payment_utr, payment_status, payment_date
            ]
            sheet.append(row_data)

            

            # Add hyperlink to the file path
            if problem.solution_pdf:
                cell = sheet.cell(row=sheet.max_row, column=len(headers) - 6)  # File link column
                cell.value = "View File"
                cell.hyperlink = solution_url
                cell.font = Font(color="0000FF", underline="single")

            # Add hyperlink to the Github link
            if problem.github_link:
                cell = sheet.cell(row=sheet.max_row, column=len(headers) - 5)
                cell.value = "View Code"
                cell.hyperlink = problem.github_link
                cell.font = Font(color="0000FF", underline="single")
            
            if payment and payment.payment_screenshot:
                cell = sheet.cell(row=sheet.max_row, column=len(headers))
                cell.value = "View Screenshot"
                cell.hyperlink = payment_screenshot
                cell.font = Font(color="0000FF", underline="single")

        # Generate the file name with timestamp
        file_name = f"submitted_teams_with_payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = f"./{file_name}"

        # Save the workbook
        workbook.save(file_path)
        self.stdout.write(self.style.SUCCESS(f"Excel file generated: {file_path}"))

